from typing import Optional, Any, Sequence, Collection, Self, Mapping, Literal
from pathlib import Path

from easyxl.exceptions import InvalidFile, InvalidSheet, InvalidRangeFormat

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.cell.cell import Cell
from openpyxl.styles.alignment import Alignment
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.utils.cell import (
    coordinate_from_string,
    column_index_from_string,
)


Coordinate = tuple[int, int]
TableData = Sequence[Sequence[Any]] | Sequence[dict[str, Any]]


class ExcelRange:
    def __init__(
        self,
        worksheet: Worksheet,
        range_expression: Optional[str] = None,
        coordinates: Optional[Sequence[tuple[int, int]]] = None,
    ) -> None:
        self.ws = worksheet
        if range_expression is not None:
            try:
                self._range = CellRange(range_expression)
            except (ValueError, TypeError):
                raise InvalidRangeFormat(
                    f"{range_expression} is not a valid range expression"
                )
        elif coordinates is not None:
            try:
                self._range = CellRange(
                    min_row=coordinates[0][0],
                    min_col=coordinates[0][1],
                    max_row=coordinates[1][0],
                    max_col=coordinates[1][1],
                )
            except ValueError | TypeError:
                raise InvalidRangeFormat(
                    f"{coordinates[0]}, {coordinates[1]} are invalid coordinate pairs"
                )
        else:
            raise InvalidRangeFormat("No expression or coordinates given")

    def __str__(self) -> str:
        return self._range.__str__()

    @property
    def row_bounds(self) -> tuple[int, int]:
        bounds = self._range.bounds
        return bounds[1], bounds[3]

    @property
    def column_bounds(self) -> tuple[str, str]:
        range_parts = self._range.__str__().split(":")
        return (
            coordinate_from_string(range_parts[0])[0],
            coordinate_from_string(range_parts[-1])[0],
        )

    @property
    def rows(self) -> Sequence[Self]:
        range_rows: list[Self] = []
        for row in self._range.rows:
            range_rows.append(type(self)(self.ws, coordinates=(row[0], row[-1])))
        return range_rows

    @property
    def next_row(self) -> Self:
        bottom_row = type(self)(
            self.ws, coordinates=(self._range.bottom[0], self._range.bottom[1])
        )
        bottom_row._range.shift(row_shift=1)
        return bottom_row

    @property
    def columns(self) -> list[Self]:
        range_columns: list[Self] = []
        for col in self._range.cols:
            range_columns.append(type(self)(self.ws, coordinates=(col[0], col[-1])))
        return range_columns

    def get_column_from_index(self, index: int | str) -> Self:
        def column_from_letter(letter: str) -> Self:
            ws_column_index = column_index_from_string(letter)
            col_start_letter, _ = self.column_bounds
            col_start = column_index_from_string(col_start_letter)
            return self.columns[ws_column_index - col_start]

        if type(index) == int:
            return self.columns[index]
        elif type(index) == str and index.isalpha():
            return column_from_letter(index)
        elif type(index) == str and index.isalnum():
            col_letter = coordinate_from_string(index)[0]
            return column_from_letter(col_letter)
        else:
            raise Exception(f"Supplied index '{index}' is not valid")

    @property
    def column_letters(self) -> list[str]:
        return [cell.column_letter for cell in self.rows[0].cells]

    @property
    def cells(self) -> list[Cell]:
        return [self.ws.cell(*cell) for cell in self._range.cells]  # type: ignore

    @property
    def cells_matrix(self) -> list[list[Cell]]:
        matrix: list[list[Cell]] = []
        for row in self.rows:
            matrix.append([c for c in row.cells])
        return matrix

    @property
    def values(self) -> list[list]:
        data: list[list] = []
        for row in self.rows:
            data.append([c.value for c in row.cells])
        return data

    @property
    def first_free_row(self) -> Self | None:
        for row in self.rows:
            if row.is_empty():
                return row
        else:
            return None

    @property
    def last_free_block(self) -> Self | None:
        free_rows_reversed: list[ExcelRange] = []
        for row in reversed(self.rows):
            if row.is_empty():
                free_rows_reversed.append(row)

        if not free_rows_reversed:
            return None

        top_left = free_rows_reversed[-1].cells[0].coordinate
        bottom_right = free_rows_reversed[0].cells[-1].coordinate
        return type(self)(self.ws, f"{top_left}:{bottom_right}")

    def is_empty(self) -> bool:
        for cell in self.cells:
            if type(cell.value) == str and cell.value.strip():  # type: ignore
                return False
            elif cell.value:
                return False
        else:
            return True

    def expand(self, right: int = 0, down: int = 0, left: int = 0, up: int = 0) -> str:
        self._range.expand(right, down, left, up)
        return self.__str__()

    def issubset(self, other: "ExcelRange") -> bool:
        return self._range.issubset(other._range)

    def issuperset(self, other: "ExcelRange") -> bool:
        return self._range.issuperset(other._range)

    def contains_cells(self, cell: Cell) -> bool:
        return self._range.issuperset(CellRange(cell.coordinate))


class ExcelRangeWritable(ExcelRange):
    @classmethod
    def convert_range_to_writable(cls, range: ExcelRange) -> "ExcelRangeWritable":
        return ExcelRangeWritable(range.ws, str(range))

    def write_to_row(self, index: int, data: Collection[str]) -> None:
        try:
            row = self.rows[index]
        except IndexError:
            raise Exception(
                f"Tried writing to row '{index}' of {self}"
                + " but that row is outside of the range."
            )

        if len(row.cells) < len(data):
            raise Exception(
                f"Row size of range {self} is '{len(row.cells)}'"
                + f" but input data is size '{len(data)}'"
            )

        for i, item in enumerate(data):
            row.cells[i].value = item

    def write_to_first_empty_row(self, data: Collection[str]) -> None:
        for i, row in enumerate(self.rows):
            if row.is_empty():
                self.write_to_row(i, data)
                break
        else:
            raise Exception(f"Range {self} has no free rows to write to.")

    def write_data(self, data: Collection[Collection[str]]) -> None:
        if len(data) > len(self.rows):
            raise Exception(
                f"Input data rows ({len(data)}) "
                + f"are greater than range rows ({self}, {len(self.rows)})"
            )

        for i, row_data in enumerate(data):
            self.write_to_row(i, row_data)

    def append_data(self, data: Collection[Collection[str]]) -> None:
        free_range = self.last_free_block
        if free_range is None:
            raise Exception(f"Range {self} has no free blocks to write lines to.")

        writeable_range = ExcelRangeWritable.convert_range_to_writable(free_range)
        writeable_range.write_data(data)

    def set_column_width(self, index: int | str, width: int) -> None:
        column = self.get_column_from_index(index)
        col_letter = column.cells[0].column_letter

        self.ws.column_dimensions[col_letter].width = width

    def auto_adjust_column_width(self, index: int | str) -> None:
        column = self.get_column_from_index(index)
        col_letter = column.cells[0].column_letter

        max_string_length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column.cells
        )
        if max_string_length > 0:
            self.ws.column_dimensions[col_letter].width = max_string_length

    def auto_adjust_column_widths(self) -> None:
        for col_index in range(len(self.columns)):
            self.auto_adjust_column_width(col_index)

    def set_alignment(
        self,
        alignment: Literal[
            "fill",
            "general",
            "justify",
            "center",
            "left",
            "centerContinuous",
            "distributed",
            "right",
        ],
    ) -> None:
        for cell in self.cells:
            cell.alignment = Alignment(horizontal=alignment)


SupportsRange = str | ExcelRange | ExcelRangeWritable


def get_range_object(
    range: SupportsRange, worksheet: Worksheet | None = None, writable: bool = False
) -> ExcelRange | ExcelRangeWritable:
    if isinstance(range, ExcelRange):
        if writable:
            return ExcelRangeWritable.convert_range_to_writable(range)
        else:
            return range

    elif type(range) == str and worksheet is not None:
        if writable:
            return ExcelRangeWritable(worksheet, range_expression=range)
        else:
            return ExcelRange(worksheet, range_expression=range)

    else:
        raise Exception(f"No valid range object for type {type(range)}")


class ExcelTable:
    def __init__(self, ws: Worksheet, table_object: Table) -> None:
        self._parent_ws = ws
        self._table = table_object
        # self.categories: list[str] = [col.name for col in self._table.tableColumns]

    def __str__(self) -> str:
        if self.name is None:
            return str(self.range)
        else:
            return f"{self.range} ({self.name})"

    @property
    def categories(self) -> list[str]:
        return [col.name for col in self._table.tableColumns]

    @property
    def columns(self) -> Mapping[str, ExcelRange]:
        return {
            str(tc.name): col
            for col, tc in zip(self.range.columns, self._table.tableColumns)
        }

    @property
    def name(self) -> str | None:
        return self._table.name

    @property
    def range(self) -> ExcelRange:
        return ExcelRange(self._parent_ws, self._table.ref)

    @property
    def first_free_row(self) -> ExcelRange:
        for row in self.range.rows[1:]:
            if row.is_empty():
                return row
        else:
            return self.range.next_row


class ExcelTableWritable(ExcelTable):
    @property
    def range(self) -> ExcelRangeWritable:
        return ExcelRangeWritable.convert_range_to_writable(super().range)

    @property
    def columns(self) -> Mapping[str, ExcelRangeWritable]:
        return {
            k: ExcelRangeWritable.convert_range_to_writable(r)
            for k, r in super().columns.items()
        }

    def append(self, data: TableData):
        category_index_map = {cat: i for i, cat in enumerate(self.categories)}

        for entry in data:
            if type(entry) == dict:
                row_values: Sequence[Any] = [None] * len(self.categories)
                for category, value in entry.items():
                    row_values[category_index_map[category]] = value

            elif isinstance(entry, Sequence):
                row_values = entry

            else:
                raise Exception("ERROR 10: YOU SHOULD NEVER SEE THIS")

            self.range.write_to_first_empty_row(row_values)

    def change_style(self, style: str) -> None:
        table_style = TableStyleInfo(name=style)
        self._table.tableStyleInfo = table_style

    def set_column_width(self, category: str, width: int) -> None:
        if (column := self.columns.get(category, None)) is not None:
            column.set_column_width(0, width)
        else:
            raise Exception(f"No column category named '{category}'")

    def auto_fit_column_widths(
        self, categories: Optional[Collection[str]] = None
    ) -> None:
        if categories is None:
            categories = self.categories

        for category in categories:
            column = self.columns[category]
            column.auto_adjust_column_widths()


class NewExcelTable(ExcelTableWritable):
    TABLE_INDEX = 0

    def __init__(
        self,
        ws: Worksheet,
        range: SupportsRange,
        categories: Optional[Collection[str]] = None,
        name: Optional[str] = None,
        initial_data: Optional[TableData] = None,
        auto_adjust_widths: bool = False,
        **table_kwargs,
    ) -> None:
        range = get_range_object(range, writable=True)
        assert isinstance(range, ExcelRangeWritable)

        if categories is not None:
            range.expand(down=-1)
            range.write_to_row(0, categories)

        table_name = name if name else self._next_table_name()

        table = Table(
            ref=str(range),
            displayName=table_name,
            **table_kwargs,
        )

        if categories is not None:
            table._initialise_columns()  # type: ignore
            for column, category in zip(table.tableColumns, categories):
                column.name = category

        ws.add_table(table)

        super().__init__(ws, ws.tables[table_name])

        if initial_data:
            self.append(initial_data)

        if auto_adjust_widths:
            range.auto_adjust_column_widths()

    def _next_table_name(self) -> str:
        NewExcelTable.TABLE_INDEX += 1
        return f"EasyXLTable{NewExcelTable.TABLE_INDEX}"


# todo: refactor so that the base classes below are in their respective files
# I thought we may write operation classes for different datatypes in Excel,
#  but it was much easier to write single operation classes for everything


class WorkbookOpenBase:
    def __init__(self, file_path: Path) -> None:
        if not file_path.is_file():
            raise FileNotFoundError(str(file_path))
        elif not file_path.suffix.startswith(".xl"):
            raise InvalidFile(file_path)

        self.file_path = file_path
        self.wb: Workbook = load_workbook(str(self.file_path))
        self.current_worksheet = self.wb.worksheets[0]

    def get_worksheet_by_title(self, title: str) -> Worksheet:
        if title not in self.wb.sheetnames:
            raise InvalidSheet(self.wb, name=title)

        return self.wb[title]

    def get_worksheet_by_index(self, index: int) -> Worksheet | None:
        if index >= len(self.wb.worksheets):
            raise InvalidSheet(self.wb, index=index)

        return self.wb.worksheets[index]


class WorkbookEditorBase(WorkbookOpenBase):
    def save(self) -> None:
        self.wb.save(self.file_path)


class WorkbookCreatorBase(WorkbookEditorBase):
    def __init__(self) -> None:
        self.wb: Workbook = Workbook(write_only=True)
        self.current_worksheet = self.wb.worksheets[0]

    def save(self, file_path: Path) -> None:
        self.wb.save(file_path)
